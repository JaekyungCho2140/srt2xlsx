"""
Excel 파일 생성 모듈
PRD 섹션 3.2 참조
"""

from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from src.srt_parser import Subtitle


class ExcelGenerator:
    """Excel 파일 생성 클래스"""

    # 헤더 정의 (PRD 섹션 3.2)
    HEADERS = [
        "#", "START", "END",
        "KO/한국어", "EN/영어", "CT/중국어 번체", "CS/중국어 간체", "JA/일본어", "TH/태국어",
 "ES-LATAM/스페인어(남미)", "PT-BR/포르투갈어(브라질)", "RU/러시아어"
    ]

    # 언어 코드 인덱스 매핑
    LANGUAGE_COLUMNS = {
        "KO": 3,
        "EN": 4,
        "CT": 5,
        "CS": 6,
        "JA": 7,
        "TH": 8,
        "ES-LATAM": 9,
        "PT-BR": 10,
        "RU": 11
    }

    @staticmethod
    def create_workbook(subtitles_by_language: Dict[str, List[Subtitle]]) -> Workbook:
        """
        Excel 워크북 생성

        Args:
            subtitles_by_language: 언어별 자막 리스트 딕셔너리
                예: {"KO": [Subtitle, ...], "EN": [Subtitle, ...]}

        Returns:
            생성된 Workbook 객체
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Subtitles"

        # 헤더 행 작성
        ExcelGenerator._write_header(ws)

        # 자막 개수 확인 (모든 언어가 동일한 개수를 가져야 함)
        subtitle_count = 0
        for lang_code, subtitles in subtitles_by_language.items():
            if subtitle_count == 0:
                subtitle_count = len(subtitles)
            elif len(subtitles) != subtitle_count:
                raise ValueError(f"언어별 자막 개수가 다릅니다: {lang_code}")

        # 자막 데이터 작성
        for row_idx in range(subtitle_count):
            # 첫 번째 언어의 자막에서 기본 정보 가져오기
            first_lang = list(subtitles_by_language.keys())[0]
            first_subtitle = subtitles_by_language[first_lang][row_idx]

            # No, Start, End 열
            ws.cell(row=row_idx + 2, column=1, value=first_subtitle.number)
            ws.cell(row=row_idx + 2, column=2, value=first_subtitle.start_time)
            ws.cell(row=row_idx + 2, column=3, value=first_subtitle.end_time)

            # 언어별 자막 텍스트
            for lang_code, subtitles in subtitles_by_language.items():
                col_idx = ExcelGenerator.LANGUAGE_COLUMNS[lang_code]
                subtitle = subtitles[row_idx]
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=subtitle.text)

        # 포매팅 적용
        ExcelGenerator._apply_formatting(ws, subtitle_count)

        return wb

    @staticmethod
    def _write_header(ws) -> None:
        """헤더 행 작성"""
        for col_idx, header in enumerate(ExcelGenerator.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)

    @staticmethod
    def _apply_formatting(ws, row_count: int) -> None:
        """
        Excel 포매팅 적용

        Args:
            ws: 워크시트 객체
            row_count: 자막 개수
        """
        # 헤더 포매팅 (굵은 글씨, 배경색 #FFF2CC, 폰트 크기 12)
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # 데이터 행 폰트 (크기 10)
        data_font = Font(size=10)

        # 셀 테두리 (얇은 검은색)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 모든 셀에 테두리, 폰트, 정렬 적용
        for row_idx in range(1, row_count + 2):
            for col_idx in range(1, len(ExcelGenerator.HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                if row_idx == 1:
                    # 헤더 행
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # 데이터 행
                    cell.font = data_font

                    # 열별 정렬 설정
                    if col_idx == 1:
                        # A열: 수직 중앙
                        cell.alignment = Alignment(vertical='center')
                    elif col_idx in [2, 3]:
                        # B, C열: 수평 중앙, 수직 중앙
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx >= 4:
                        # D~L열: 수직 중앙, 자동 줄 바꿈 해제
                        cell.alignment = Alignment(vertical='center', wrap_text=False)

        # 열 너비 설정
        # A열: 너비 4
        ws.column_dimensions['A'].width = 4

        # B, C열: 너비 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12

        # D~L열: 너비 24
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col_letter].width = 24

    @staticmethod
    def read_workbook(filepath: str) -> Dict[str, List[Subtitle]]:
        """
        Excel 파일 읽기

        Args:
            filepath: Excel 파일 경로

        Returns:
            언어별 자막 리스트 딕셔너리
        """
        from openpyxl import load_workbook
        from src.errors import ExcelHeaderError

        # data_only=True: 수식 대신 캐시된 값 읽기 (방어적 구현)
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        # 헤더 검증
        headers = []
        for col_idx in range(1, len(ExcelGenerator.HEADERS) + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            headers.append(cell_value)

        if headers != ExcelGenerator.HEADERS:
            raise ExcelHeaderError()

        # 자막 데이터 읽기
        subtitles_by_language = {}
        row_idx = 2

        while True:
            # No 열이 비어있으면 종료
            no_value = ws.cell(row=row_idx, column=1).value
            if no_value is None:
                break

            # 자막 번호, 시작 시간, 종료 시간
            number = int(no_value)
            start_time = ws.cell(row=row_idx, column=2).value
            end_time = ws.cell(row=row_idx, column=3).value

            # 언어별 텍스트 읽기
            for lang_code, col_idx in ExcelGenerator.LANGUAGE_COLUMNS.items():
                text = ws.cell(row=row_idx, column=col_idx + 1).value

                # None을 빈 문자열로 처리
                if text is None:
                    text = ""
                else:
                    text = str(text)

                # Subtitle 객체 생성
                subtitle = Subtitle(
                    number=number,
                    start_time=start_time,
                    end_time=end_time,
                    text=text
                )

                # 언어별 리스트에 추가
                if lang_code not in subtitles_by_language:
                    subtitles_by_language[lang_code] = []
                subtitles_by_language[lang_code].append(subtitle)

            row_idx += 1

        return subtitles_by_language

    @staticmethod
    def is_language_column_empty(subtitles: List[Subtitle]) -> bool:
        """
        언어 열이 비어있는지 확인

        Args:
            subtitles: 자막 리스트

        Returns:
            모든 텍스트가 비어있으면 True, 하나라도 내용이 있으면 False
        """
        for subtitle in subtitles:
            text = subtitle.text
            # None 또는 빈 문자열이 아닌 경우 (공백 제거 후 검사)
            if text and text.strip():
                return False
        return True
